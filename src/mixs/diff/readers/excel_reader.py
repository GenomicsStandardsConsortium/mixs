"""Excel reader for .xls and .xlsx MIxS schema files.

Supports MIxS versions 2-5 which were distributed as Excel spreadsheets.
Format-specific parsing rules are loaded from YAML profiles in:
    assets/legacy_format_profiles/

This allows version-specific column mappings and sheet names to be
maintained without code changes.
"""

import logging
from pathlib import Path
from typing import Dict, List, Optional, Any

import yaml

from mixs.diff.models import NormalizedSchema, NormalizedTerm
from mixs.diff.readers.base import BaseReader, detect_format

logger = logging.getLogger(__name__)

# Default location for format profiles
DEFAULT_PROFILES_DIR = Path(__file__).parent.parent.parent.parent.parent / "assets" / "legacy_format_profiles"


class FormatProfile:
    """Loaded format profile for Excel parsing."""

    def __init__(self, profile_data: Dict[str, Any]):
        self.description = profile_data.get("description", "")
        self.version_patterns = profile_data.get("version_patterns", [])
        self.sheet_names = profile_data.get("sheet_names", {})
        self.column_mappings = profile_data.get("column_mappings", {})
        self.checklist_columns = profile_data.get("checklist_columns", {})
        self.normalize_term_names = profile_data.get("normalize_term_names", False)
        self.notes = profile_data.get("notes", "")

    @classmethod
    def load(cls, path: Path) -> "FormatProfile":
        """Load profile from YAML file."""
        with open(path, 'r') as f:
            data = yaml.safe_load(f)
        return cls(data)

    def matches(self, filename: str, filepath: str) -> bool:
        """Check if this profile matches the given file."""
        combined = f"{filepath}/{filename}".lower()
        return any(pattern.lower() in combined for pattern in self.version_patterns)


class ProfileManager:
    """Manages loading and selecting format profiles."""

    def __init__(self, profiles_dir: Optional[Path] = None):
        self.profiles_dir = profiles_dir or DEFAULT_PROFILES_DIR
        self._profiles: List[FormatProfile] = []
        self._default_profile: Optional[FormatProfile] = None
        self._loaded = False

    def _ensure_loaded(self):
        """Lazy-load profiles on first access."""
        if self._loaded:
            return

        if not self.profiles_dir.exists():
            logger.warning(f"Profiles directory not found: {self.profiles_dir}")
            self._loaded = True
            return

        for profile_path in self.profiles_dir.glob("*.yaml"):
            try:
                profile = FormatProfile.load(profile_path)
                if profile_path.name == "excel_default.yaml":
                    self._default_profile = profile
                else:
                    self._profiles.append(profile)
                logger.debug(f"Loaded profile: {profile_path.name}")
            except Exception as e:
                logger.warning(f"Failed to load profile {profile_path}: {e}")

        self._loaded = True

    def get_profile(self, filepath: str) -> FormatProfile:
        """Get the best matching profile for a file."""
        self._ensure_loaded()

        filename = Path(filepath).name

        # Try to match specific profiles first
        for profile in self._profiles:
            if profile.matches(filename, filepath):
                logger.info(f"Using profile: {profile.description}")
                return profile

        # Fall back to default profile
        if self._default_profile:
            logger.info(f"Using default profile for: {filename}")
            return self._default_profile

        # Emergency fallback - return empty profile
        logger.warning(f"No profile found for {filename}, using empty profile")
        return FormatProfile({})


# Global profile manager instance
_profile_manager: Optional[ProfileManager] = None


def get_profile_manager(profiles_dir: Optional[Path] = None) -> ProfileManager:
    """Get or create the profile manager."""
    global _profile_manager
    if _profile_manager is None or profiles_dir is not None:
        _profile_manager = ProfileManager(profiles_dir)
    return _profile_manager


class ExcelReader(BaseReader):
    """Reader for Excel-based MIxS schemas (.xls and .xlsx)."""

    def __init__(self, profiles_dir: Optional[Path] = None):
        """Initialize reader with optional custom profiles directory."""
        self.profile_manager = get_profile_manager(profiles_dir)

    @classmethod
    def can_read(cls, path: str) -> bool:
        """Check if this reader can handle the given path."""
        fmt = detect_format(path)
        return fmt in ("xlsx", "xls")

    def read(self, path: str, version: Optional[str] = None) -> NormalizedSchema:
        """Read Excel schema file and return normalized representation.

        Args:
            path: Path to the Excel file.
            version: Optional version string. If not provided, attempts to detect.

        Returns:
            NormalizedSchema with terms extracted from the Excel file.
        """
        fmt = detect_format(path)
        file_path = Path(path)

        if not file_path.exists():
            raise FileNotFoundError(f"Excel file not found: {path}")

        # Load appropriate format profile
        profile = self.profile_manager.get_profile(path)

        # Detect version from path if not provided
        if version is None:
            version = self._detect_version(path)

        logger.info(f"Reading {fmt} file: {path} (version: {version})")

        if fmt == "xlsx":
            return self._read_xlsx(file_path, version, profile)
        elif fmt == "xls":
            return self._read_xls(file_path, version, profile)
        else:
            raise ValueError(f"Unsupported Excel format: {fmt}")

    def _detect_version(self, path: str) -> str:
        """Attempt to detect MIxS version from file path."""
        path_lower = path.lower()
        if "mixs5" in path_lower or "v5" in path_lower or "20180621" in path:
            return "v5"
        elif "mixs4" in path_lower or "v4" in path_lower or "210514" in path:
            return "v4"
        elif "2011" in path:
            return "v2.1"
        elif "2010" in path:
            return "v2.0"
        elif "2009" in path:
            return "v2.0"
        else:
            return "unknown"

    def _read_xlsx(self, file_path: Path, version: str, profile: FormatProfile) -> NormalizedSchema:
        """Read .xlsx file using openpyxl."""
        try:
            import openpyxl
        except ImportError:
            raise ImportError(
                "openpyxl is required for reading .xlsx files. "
                "Install with: poetry install --with legacy-diff"
            )

        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)

        schema = NormalizedSchema(
            version=version,
            source_path=str(file_path),
            source_format="xlsx",
        )

        # Find main terms sheet using profile
        main_sheet = self._find_sheet(wb.sheetnames, profile.sheet_names.get("main_terms", []))
        if main_sheet:
            self._process_main_sheet_xlsx(wb[main_sheet], schema, profile)

        # Find packages sheet using profile
        pkg_sheet = self._find_sheet(wb.sheetnames, profile.sheet_names.get("packages", []))
        if pkg_sheet:
            self._process_packages_sheet_xlsx(wb[pkg_sheet], schema, profile)

        wb.close()

        logger.info(f"Loaded {len(schema.terms)} terms, {len(schema.packages)} packages, "
                   f"{len(schema.checklists)} checklists from {file_path.name}")

        return schema

    def _read_xls(self, file_path: Path, version: str, profile: FormatProfile) -> NormalizedSchema:
        """Read .xls file using xlrd."""
        try:
            import xlrd
        except ImportError:
            raise ImportError(
                "xlrd is required for reading .xls files. "
                "Install with: poetry install --with legacy-diff"
            )

        wb = xlrd.open_workbook(file_path)

        schema = NormalizedSchema(
            version=version,
            source_path=str(file_path),
            source_format="xls",
        )

        # Find main terms sheet using profile
        main_sheet = self._find_sheet(wb.sheet_names(), profile.sheet_names.get("main_terms", []))
        if main_sheet:
            self._process_main_sheet_xls(wb.sheet_by_name(main_sheet), schema, profile)

        # Find packages sheet using profile
        pkg_sheet = self._find_sheet(wb.sheet_names(), profile.sheet_names.get("packages", []))
        if pkg_sheet:
            self._process_packages_sheet_xls(wb.sheet_by_name(pkg_sheet), schema, profile)

        logger.info(f"Loaded {len(schema.terms)} terms, {len(schema.packages)} packages, "
                   f"{len(schema.checklists)} checklists from {file_path.name}")

        return schema

    def _find_sheet(self, available_sheets: List[str], preferred_names: List[str]) -> Optional[str]:
        """Find the first matching sheet name."""
        for name in preferred_names:
            if name in available_sheets:
                return name
        return None

    def _process_main_sheet_xlsx(self, sheet, schema: NormalizedSchema, profile: FormatProfile) -> None:
        """Process the main terms sheet from an xlsx file."""
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            return

        headers = [str(h).strip() if h else "" for h in rows[0]]
        col_map = self._build_column_map(headers, profile)
        checklist_cols = self._find_checklist_columns(headers, profile)

        # Initialize checklists
        for checklist_name in checklist_cols.values():
            schema.checklists[checklist_name] = []

        for row in rows[1:]:
            if not row or not row[0]:
                continue

            term = self._row_to_term(row, col_map, checklist_cols, profile)
            if term and term.name:
                schema.terms[term.name] = term

                # Add to checklists based on membership
                for checklist_name, requirement in term.checklist_membership.items():
                    if requirement and requirement.strip():
                        if checklist_name not in schema.checklists:
                            schema.checklists[checklist_name] = []
                        schema.checklists[checklist_name].append(term.name)

    def _process_main_sheet_xls(self, sheet, schema: NormalizedSchema, profile: FormatProfile) -> None:
        """Process the main terms sheet from an xls file."""
        if sheet.nrows == 0:
            return

        headers = [str(sheet.cell_value(0, c)).strip() for c in range(sheet.ncols)]
        col_map = self._build_column_map(headers, profile)
        checklist_cols = self._find_checklist_columns(headers, profile)

        # Initialize checklists
        for checklist_name in checklist_cols.values():
            schema.checklists[checklist_name] = []

        for row_idx in range(1, sheet.nrows):
            row = [sheet.cell_value(row_idx, c) for c in range(sheet.ncols)]
            if not row or not row[0]:
                continue

            term = self._row_to_term(row, col_map, checklist_cols, profile)
            if term and term.name:
                schema.terms[term.name] = term

                # Add to checklists based on membership
                for checklist_name, requirement in term.checklist_membership.items():
                    if requirement and requirement.strip():
                        if checklist_name not in schema.checklists:
                            schema.checklists[checklist_name] = []
                        schema.checklists[checklist_name].append(term.name)

    def _process_packages_sheet_xlsx(self, sheet, schema: NormalizedSchema, profile: FormatProfile) -> None:
        """Process the packages sheet from xlsx."""
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            return

        headers = [str(h).strip() if h else "" for h in rows[0]]
        col_map = self._build_column_map(headers, profile)

        for row in rows[1:]:
            if not row:
                continue

            # Get package name and term name
            pkg_idx = col_map.get("package_name")
            name_idx = col_map.get("name")

            if pkg_idx is None or name_idx is None:
                continue

            pkg_name = str(row[pkg_idx]).strip() if row[pkg_idx] else ""
            term_name = str(row[name_idx]).strip() if row[name_idx] else ""

            if not pkg_name or not term_name:
                continue

            # Normalize names
            pkg_name_normalized = self._normalize_package_name(pkg_name)
            if profile.normalize_term_names:
                term_name = self._normalize_term_name(term_name)

            if pkg_name_normalized not in schema.packages:
                schema.packages[pkg_name_normalized] = []
            schema.packages[pkg_name_normalized].append(term_name)

            # Update term's package membership if term exists
            if term_name in schema.terms:
                req_idx = col_map.get("requirement")
                requirement = str(row[req_idx]).strip() if req_idx and row[req_idx] else ""
                schema.terms[term_name].package_membership[pkg_name_normalized] = requirement

    def _process_packages_sheet_xls(self, sheet, schema: NormalizedSchema, profile: FormatProfile) -> None:
        """Process the packages sheet from xls."""
        if sheet.nrows == 0:
            return

        headers = [str(sheet.cell_value(0, c)).strip() for c in range(sheet.ncols)]
        col_map = self._build_column_map(headers, profile)

        for row_idx in range(1, sheet.nrows):
            row = [sheet.cell_value(row_idx, c) for c in range(sheet.ncols)]
            if not row:
                continue

            pkg_idx = col_map.get("package_name")
            name_idx = col_map.get("name")

            if pkg_idx is None or name_idx is None:
                continue

            pkg_name = str(row[pkg_idx]).strip() if row[pkg_idx] else ""
            term_name = str(row[name_idx]).strip() if row[name_idx] else ""

            if not pkg_name or not term_name:
                continue

            pkg_name_normalized = self._normalize_package_name(pkg_name)
            if profile.normalize_term_names:
                term_name = self._normalize_term_name(term_name)

            if pkg_name_normalized not in schema.packages:
                schema.packages[pkg_name_normalized] = []
            schema.packages[pkg_name_normalized].append(term_name)

            if term_name in schema.terms:
                req_idx = col_map.get("requirement")
                requirement = str(row[req_idx]).strip() if req_idx and row[req_idx] else ""
                schema.terms[term_name].package_membership[pkg_name_normalized] = requirement

    def _build_column_map(self, headers: List[str], profile: FormatProfile) -> Dict[str, int]:
        """Build mapping from normalized field names to column indices using profile."""
        col_map = {}
        for field_name, variations in profile.column_mappings.items():
            for i, header in enumerate(headers):
                if header in variations:
                    col_map[field_name] = i
                    break
        return col_map

    def _find_checklist_columns(self, headers: List[str], profile: FormatProfile) -> Dict[int, str]:
        """Find checklist columns using profile mappings.

        Returns:
            Dict mapping column index to normalized checklist name.
        """
        checklist_cols = {}
        for i, header in enumerate(headers):
            header_stripped = header.strip()
            if header_stripped in profile.checklist_columns:
                checklist_cols[i] = profile.checklist_columns[header_stripped]
        return checklist_cols

    def _row_to_term(self, row: List[Any], col_map: Dict[str, int],
                     checklist_cols: Dict[int, str], profile: FormatProfile) -> Optional[NormalizedTerm]:
        """Convert a row to a NormalizedTerm."""
        def get_val(field: str) -> str:
            idx = col_map.get(field)
            if idx is not None and idx < len(row) and row[idx] is not None:
                return str(row[idx]).strip()
            return ""

        name = get_val("name")
        if not name:
            return None

        # Normalize the name if profile requires it
        if profile.normalize_term_names:
            name = self._normalize_term_name(name)

        # For older formats, "description" may contain expected value hints
        expected_value = get_val("expected_value")
        if not expected_value:
            expected_value = get_val("description")

        term = NormalizedTerm(
            name=name,
            item=get_val("item") or get_val("name"),
            definition=get_val("definition"),
            example=get_val("example"),
            expected_value=expected_value,
            value_syntax=get_val("value_syntax"),
            occurrence=get_val("occurrence"),
            preferred_unit=get_val("preferred_unit"),
            section=get_val("section"),
            position=get_val("position"),
            mixs_id=get_val("mixs_id"),
        )

        # Extract checklist membership
        for col_idx, checklist_name in checklist_cols.items():
            if col_idx < len(row) and row[col_idx]:
                val = str(row[col_idx]).strip()
                if val:
                    term.checklist_membership[checklist_name] = val

        return term

    def _normalize_term_name(self, name: str) -> str:
        """Normalize term name to structured comment format.

        Converts names like "Project Name" to "project_name".
        """
        normalized = name.lower()
        normalized = normalized.replace(" ", "_")
        normalized = normalized.replace("-", "_")
        while "__" in normalized:
            normalized = normalized.replace("__", "_")
        return normalized.strip("_")

    def _normalize_package_name(self, name: str) -> str:
        """Normalize package name."""
        return name.lower().replace(" ", "_").replace("-", "_")
