#!/usr/bin/env python3
"""
One-time MIxS v5 (2018 Excel) to v6.0.0 (2022 LinkML) diff.

This is NOT the reusable diff tool. For comparing any two LinkML versions of MIxS
(v6 onward), use src/scripts/diff_two_linkml_mixs_releases.py, documented in
contrib/docs/SCHEMA_DIFFING.md.

Both versions here are finished releases, pinned by commit and tag, so the script
takes no arguments and the result never changes. It writes a structured diff and
a short summary to assets/diff_results/v5_to_v6.0.0/.

Needs openpyxl to read the v5 Excel file (installed by `poetry install`).

How it works:
- The v5 Excel spreads its 600 terms over two sheets: "MIxS" (94 core terms) and
  "environmental_packages" (509 more, each repeated once per package). We read
  both and deduplicate by term name.
- The v6.0.0 side is read from the mixs6.0.0 tag with LinkML's SchemaView, using
  induced slots so inherited and mixin values are resolved.
- A small, checked rename and deletion map (below) accounts for the names that
  changed or were removed between v5 and v6.0.0.
"""
import difflib
import os
import subprocess
import sys
import tempfile
import urllib.request
from pathlib import Path

# Fix the hash seed so set iteration is stable and re-runs write the same files.
if os.environ.get("PYTHONHASHSEED") != "0":
    os.environ["PYTHONHASHSEED"] = "0"
    os.execv(sys.executable, [sys.executable, *sys.argv])

import yaml
from openpyxl import load_workbook
from linkml_runtime.utils.schemaview import SchemaView

HERE = Path(__file__).resolve().parent
REPO_ROOT = HERE.parent.parent  # src/scripts -> src -> repo root
OUT_DIR = REPO_ROOT / "assets" / "diff_results" / "v5_to_v6.0.0"

# Pinned inputs (finished releases).
V5_XLSX_URL = (
    "https://raw.githubusercontent.com/GenomicsStandardsConsortium/mixs-legacy/"
    "9628b5e5aa89/mixs5/mixs_v5.xlsx"
)
V6_TAG = "mixs6.0.0"
V6_SCHEMA = "model/schema/mixs.yaml"

# v5 -> v6.0.0 name changes, each target checked to exist in the mixs6.0.0 tag.
# Several are abbreviations introduced at v6.0.0 (for example content -> cont,
# collection -> collect, dropped host_/ihmc_ prefixes). The rename_candidates
# section in the output flags any further removed/added pairs that look like
# renames but are not listed here, so missed renames surface on the next run.
RENAMES = {
    "16s_recover": "x_16s_recover",
    "16s_recover_software": "x_16s_recover_software",
    "health_disease_stat": "host_disease_stat",
    "votu_class_appr": "otu_class_appr",
    "votu_db": "otu_db",
    "votu_seq_comp_appr": "otu_seq_comp_appr",
    "chem_treatment_method": "chem_treat_method",
    "heat_system_deliv_meth": "heat_sys_deliv_meth",
    "host_blood_press_diast": "blood_press_diast",
    "host_blood_press_syst": "blood_press_syst",
    "ihmc_ethnicity": "ethnicity",
    "organism_count_qpcr_info": "org_count_qpcr_info",
    "room_architec_element": "room_architec_elem",
    "room_moist_damage_hist": "room_moist_dam_hist",
    "samp_collection_point": "samp_collect_point",
    "shading_device_water_mold": "shad_dev_water_mold",
    "tot_nitro_content_meth": "tot_nitro_cont_meth",
    "tvdss_of_hcr_pressure": "tvdss_of_hcr_press",
    "water_content_soil_meth": "water_cont_soil_meth",
    "water_production_rate": "water_prod_rate",
}
# Terms removed on purpose at v6.0.0 (structural), not counted as unexplained.
DELETIONS = {"env_package", "investigation_type", "submitted_to_insdc"}


def read_v5_terms() -> dict:
    """Return {term_name: definition} from the two term sheets of the v5 Excel."""
    tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    urllib.request.urlretrieve(V5_XLSX_URL, tmp.name)
    wb = load_workbook(tmp.name, read_only=True, data_only=True)
    terms: dict = {}
    for sheet in ("MIxS", "environmental_packages"):
        rows = wb[sheet].iter_rows(values_only=True)
        header = list(next(rows))
        name_i = header.index("Structured comment name")
        def_i = header.index("Definition")
        for row in rows:
            name = row[name_i]
            if not name:
                continue
            terms.setdefault(str(name).strip(), (row[def_i] or "") if def_i < len(row) else "")
    wb.close()
    os.unlink(tmp.name)
    return terms


def read_v6_terms() -> dict:
    """Return {slot_name: description} from the mixs6.0.0 tag (induced slots)."""
    tmp = tempfile.mkdtemp(prefix="mixs600_")
    with open(Path(tmp) / "schema.tar", "wb") as out:
        subprocess.run(["git", "archive", V6_TAG, "model/schema"],
                       check=True, stdout=out, cwd=REPO_ROOT)
    subprocess.run(["tar", "-xf", "schema.tar"], check=True, cwd=tmp)
    sv = SchemaView(str(Path(tmp) / V6_SCHEMA))
    return {name: (sv.induced_slot(name, imports=True).description or "")
            for name in sv.all_slots(imports=True)}


def main() -> None:
    v5 = read_v5_terms()
    v6 = read_v6_terms()
    v6_names = set(v6)

    shared, renamed, removed, definition_changed = {}, {}, [], {}
    for name in sorted(v5):
        if name in DELETIONS:
            continue
        target = RENAMES.get(name, name)
        if target not in v6_names:
            removed.append(name)
            continue
        (renamed if target != name else shared)[name] = target
        if (v5[name] or "").strip() != (v6[target] or "").strip():
            definition_changed[name] = {"v5": v5[name], "v6.0.0": v6[target]}
    added = sorted(v6_names - set(shared.values()) - set(renamed.values()))

    # Self-check: a removed name that closely matches an added name is probably a
    # rename that is missing from RENAMES. Flag those for review; promote real
    # ones into RENAMES so they stop showing as a removal plus an addition.
    rename_candidates = {}
    for r in removed:
        best = max(added, default=None,
                   key=lambda a: difflib.SequenceMatcher(None, r, a).ratio())
        if best and difflib.SequenceMatcher(None, r, best).ratio() >= 0.75:
            rename_candidates[r] = best

    result = {
        "comparison": {
            "old": {"version": "v5", "source": V5_XLSX_URL, "terms": len(v5)},
            "new": {"version": "v6.0.0",
                    "source": f"GenomicsStandardsConsortium/mixs@{V6_TAG}:{V6_SCHEMA}",
                    "terms": len(v6)},
        },
        "counts": {"shared": len(shared), "renamed": len(renamed),
                   "removed": len(removed), "deleted": len(DELETIONS),
                   "added": len(added), "definition_changed": len(definition_changed),
                   "rename_candidates": len(rename_candidates)},
        "renamed": renamed,
        "rename_candidates": rename_candidates,
        "deleted": sorted(DELETIONS),
        "removed": sorted(removed),
        "added": added,
        "definition_changed": definition_changed,
    }

    OUT_DIR.mkdir(parents=True, exist_ok=True)
    (OUT_DIR / "schema_comparison_results.yaml").write_text(
        yaml.safe_dump(result, sort_keys=True, allow_unicode=True, width=100))
    c = result["counts"]
    cand = "".join(f"- {o} -> {n}\n" for o, n in sorted(rename_candidates.items()))
    (OUT_DIR / "tool_summary.md").write_text(
        f"# MIxS v5 to v6.0.0 diff\n\n"
        f"- v5 (Excel): {len(v5)} terms\n- v6.0.0 (LinkML): {len(v6)} terms\n\n"
        f"- shared: {c['shared']}\n- renamed: {c['renamed']}\n- removed: {c['removed']}\n"
        f"- deleted (structural): {c['deleted']}\n- added: {c['added']}\n"
        f"- definition changed: {c['definition_changed']}\n"
        + (f"\n## Possible missed renames (review, then add to RENAMES)\n\n{cand}"
           if rename_candidates else ""))
    print("counts:", c)
    if rename_candidates:
        print("possible missed renames:", rename_candidates)


if __name__ == "__main__":
    main()
